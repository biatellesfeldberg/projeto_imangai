"""
gerar_planilha.py
-----------------
Lê o CSV produzido pelo coletor (`saida/casas_filtradas.csv` por padrão) e gera uma planilha
Excel (.xlsx) formatada na pasta `planilhas_geradas/`.

O arquivo de saída padrão é sempre regravado por completo (substitui o .xlsx anterior se existir).

Dependência:
    pip install openpyxl

Uso:
    python gerar_planilha.py
    python gerar_planilha.py --entrada saida/casas_filtradas.csv --saida planilhas_geradas/casas.xlsx
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Cores e estilo (visual limpo, bom contraste em Excel/Numbers)
FILL_CABECALHO = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FONTE_CABECALHO = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONTE_CORPO = Font(name="Calibri", size=11, color="000000")
FONTE_LINK = Font(name="Calibri", size=11, color="0563C1", underline="single")
FONTE_INDICE = Font(name="Calibri", size=11, bold=False, color="404040")
BORDA_FINA = Side(style="thin", color="B4B4B4")
BORDA_CELULA = Border(
    left=BORDA_FINA, right=BORDA_FINA, top=BORDA_FINA, bottom=BORDA_FINA
)
FILL_ZEBRA = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")


def _ajustar_larguras(ws: object, larguras: dict[int, float]) -> None:
    for col_idx, w in larguras.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _aplicar_borda(cell: object) -> None:
    cell.border = BORDA_CELULA


def csv_para_planilha(
    caminho_csv: Path,
    caminho_xlsx: Path,
) -> int:
    if not caminho_csv.is_file():
        raise FileNotFoundError(f"Arquivo CSV não encontrado: {caminho_csv}")

    caminho_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Casas"

    cabecalhos = [
        "Índice",
        "Link do anúncio",
        "Endereço",
        "Telefone da imobiliária",
        "Telefone do vendedor",
    ]
    ws.row_dimensions[1].height = 26
    for col, titulo in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = FONTE_CABECALHO
        cell.fill = FILL_CABECALHO
        cell.alignment = Alignment(
            horizontal="center" if col == 1 else "left",
            vertical="center",
            wrap_text=True,
        )
        _aplicar_borda(cell)

    linhas = 0
    with caminho_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for idx, row in enumerate(reader, start=1):
            link = (row.get("link") or "").strip()
            endereco = (row.get("endereco") or "").strip()
            tel_imob = (row.get("telefone_imobiliaria") or "").strip()
            tel_vend = (row.get("telefone_vendedor") or "").strip()

            r = idx + 1
            ws.row_dimensions[r].height = 38

            c_idx = ws.cell(row=r, column=1, value=idx)
            c_idx.font = FONTE_INDICE
            c_idx.alignment = Alignment(horizontal="center", vertical="top")
            _aplicar_borda(c_idx)

            c_link = ws.cell(row=r, column=2, value=link)
            if link.lower().startswith(("http://", "https://")):
                c_link.hyperlink = link
                c_link.font = FONTE_LINK
            else:
                c_link.font = FONTE_CORPO
            c_link.alignment = Alignment(vertical="top", wrap_text=True)
            _aplicar_borda(c_link)

            c_end = ws.cell(row=r, column=3, value=endereco)
            c_end.font = FONTE_CORPO
            c_end.alignment = Alignment(vertical="top", wrap_text=True)
            _aplicar_borda(c_end)

            c_imob = ws.cell(row=r, column=4, value=tel_imob)
            c_imob.font = FONTE_CORPO
            c_imob.alignment = Alignment(vertical="top", wrap_text=False)
            _aplicar_borda(c_imob)

            c_vend = ws.cell(row=r, column=5, value=tel_vend if tel_vend else "")
            c_vend.font = FONTE_CORPO
            c_vend.alignment = Alignment(vertical="top", wrap_text=False)
            _aplicar_borda(c_vend)

            if idx % 2 == 0:
                for c in range(1, 6):
                    ws.cell(row=r, column=c).fill = FILL_ZEBRA

            linhas += 1

    ws.freeze_panes = "A2"
    _ajustar_larguras(
        ws,
        {
            1: 11,
            2: 54,
            3: 50,
            4: 24,
            5: 24,
        },
    )

    wb.save(caminho_xlsx)
    return linhas


def main() -> None:
    raiz = Path(__file__).resolve().parent
    ap = argparse.ArgumentParser(
        description="Gera planilha Excel legível a partir do CSV do web scraping."
    )
    ap.add_argument(
        "--entrada",
        default=str(raiz / "saida" / "casas_filtradas.csv"),
        help="CSV gerado pelo coletor (padrão: saida/casas_filtradas.csv)",
    )
    ap.add_argument(
        "--saida",
        default=str(raiz / "planilhas_geradas" / "planilha_casas.xlsx"),
        help="Arquivo .xlsx de saída (padrão: planilhas_geradas/planilha_casas.xlsx)",
    )
    args = ap.parse_args()

    csv_path = Path(args.entrada).resolve()
    xlsx_path = Path(args.saida).resolve()

    n = csv_para_planilha(csv_path, xlsx_path)
    print(f"Planilha salva em: {xlsx_path}")
    print(f"Total de casas (linhas de dados): {n}")


if __name__ == "__main__":
    main()
